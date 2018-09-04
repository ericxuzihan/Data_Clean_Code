from xlrd import open_workbook
from xlutils.copy import copy
import openpyxl


def upc_hash_dict(filename):
    rb = open_workbook(filename)
    r_sheet = rb.sheet_by_index(0)
    wb = copy(rb)
    # w_sheet = wb.get_sheet(0)
    r_sheet.cell_value(0, 0)
    first_row_list = r_sheet.row_values(0)
    UPC_index = first_row_list.index('UPC')
    FBA_price_index = first_row_list.index('FBA Sale At Current BuyBox - Net Payment')
    # print(UPC_index)  # test if the column number is expected
    # print(FBA_price_index)
    connect_dict = {}
    for sheetname in rb.sheet_names():
        oldsheet = rb.sheet_by_name(sheetname)
        for i in range(oldsheet.nrows):
            UPC_value = str(oldsheet.cell(i, UPC_index).value)
            FBA_sale_price = str(oldsheet.cell(i, FBA_price_index).value)
            connect_dict[UPC_value] = FBA_sale_price
            # print(UPC_value)
            # print(FBA_sale_price)
    del connect_dict['UPC']

    for i in connect_dict:
        # print(i[0])
        if i[0] == '0':
            # i = i[1:]
            connect_dict[i[1:]] = connect_dict[i]
            del connect_dict[i]
    print(connect_dict)

    return connect_dict, UPC_index

#######################################################################################
# ################################Driver program#######################################


connector = upc_hash_dict('C:\\Users\\Eric\\Desktop\\New folder\\price_list_306098_file_1_of_1.xlsx')[0]
UPC_index = upc_hash_dict('C:\\Users\\Eric\\Desktop\\New folder\\price_list_306098_file_1_of_1.xlsx')[1]
print(connector)
print(UPC_index)
######################################################################################
######################################################################################


def complete_file(complete_filename):
    wb = openpyxl.load_workbook(complete_filename)
    worksheet = wb['RepricerPricingTiersResults']
    # worksheet = wb.get_sheet_by_name('RepricerPricingTiersResults')
    worksheet['I1'] = 'FBA Sale At Current BuyBox - Net Payment'
    for i in connector:
        for j in range(1, 200000):
            a = worksheet.cell(row=j, column=2).value
            # print(a)
            if i == a:
                print(j)
                worksheet.cell(row=j, column=9).value = connector[i]
    wb.save('C:\\Users\\Eric\\Desktop\\Test_Data_Sheet\\repricing_output.xls')

    return


complete_file('C:\\Users\\Eric\\Desktop\\Test_Data_Sheet\\RepricerPricingTiersResults27.xlsx')


# if __name__ == "__main__":
"""""""""
def complete_file(complete_filename):
    rb = open_workbook(complete_filename)
    r_sheet = rb.sheet_by_index(0)
    wb = copy(rb)
    w_sheet = wb.get_sheet(0)
    #w_sheet.getCells().insertColumns(8, 1)

    #w_sheet.cell_value(8, 0) == "eric is awesome"
    r_sheet.cell_value(0, 0)
    w_sheet.write(0, 8, "FBA Sale At Current BuyBox - Net Payment")

    for sheetname in rb.sheet_names():
        oldsheet = rb.sheet_by_name(sheetname)
        for i in connector:
            for j in range(oldsheet.nrows):
                if i == str(oldsheet.cell(j, UPC_index).value):
                    w_sheet.write(j, 8, connector[i])
    wb.save('C:\\Users\\Eric\\Desktop\\Test_Data_Sheet\\repricing_output.xls')

    return
    
    
def complete_file(complete_filename):
    workbook = xlsxwriter.Workbook(complete_filename)
    worksheet = workbook.add_worksheet()
    #rb = open_workbook(complete_filename)
    #r_sheet = rb.sheet_by_index(0)
    #wb = copy(rb)
    #w_sheet = wb.get_sheet(0)
    #w_sheet.getCells().insertColumns(8, 1)

    #w_sheet.cell_value(8, 0) == "eric is awesome"
    #r_sheet.cell_value(0, 0)
    worksheet.write(0, 8, "FBA Sale At Current BuyBox - Net Payment")

    #for sheetname in workbook.sheet_names():
        #oldsheet = workbook.sheet_by_name(sheetname)
    for i in connector:
        for j in range(1,200000):
            if i == xl_rowcol_to_cell(j, UPC_index):
                worksheet.write(j, 8, connector[i])
    #workbook.save('C:\\Users\\Eric\\Desktop\\Test_Data_Sheet\\repricing_output.xls')

    return

"""""